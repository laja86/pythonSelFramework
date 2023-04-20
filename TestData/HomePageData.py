import openpyxl


class HomePageData:

    # test_HomePage_data = [{"firstname":"Rahul","lastname":"shetty","gender":"Male"}, {"firstname":"Anshika", "lastname":"shetty", "gender":"Female"}]

    # SELF parameter not present, because the method is declared as static. This is to avoid creating objects to call this object
    # self parameter is only require for non-static methods
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        #book = openpyxl.load_workbook("C:\\Users\\luis.juarez\\PycharmProjects\\pythonSelFramework\\TestData\\PythonDemo.xlsx")
        #sheet = book.active
        #for i in range(1, sheet.max_row + 1):  # to get rows
        #    if sheet.cell(row=i, column=1).value == test_case_name:

        #        for j in range(2, sheet.max_column + 1):  # to get columns
        #            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        #return[Dict]  # will be sent in a list
        test_HomePage_data = [{"firstname": "Rahul", "lastname": "shetty", "gender": "Male"},
                              {"firstname": "Anshika", "lastname": "shetty", "gender": "Female"}]

        return test_HomePage_data;
